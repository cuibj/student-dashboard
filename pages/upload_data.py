import streamlit as st
import pandas as pd
import openpyxl
from sqlalchemy import create_engine
import sqlite3
import datetime
from st_btn_group import st_btn_group
# from streamlit.components.v1 import html

st.title('成绩分析')

uploaded_file = st.file_uploader("Choose a file", type = 'xlsx')


@st.cache_resource
def get_conn():
    db_path = '/Users/amber/workspace/student-dashboard/data/student.db'
    conn = sqlite3.connect(db_path)
    engine = create_engine(f'sqlite:///{db_path}', echo=False)
    return conn,engine

conn,engine = get_conn()


def delete_insert(cur, data, exam_name):
    del_sql = f'''
        delete from scores 
        where 考试名称 = '{exam_name}'
    '''
    cur = conn.cursor()
    cur.execute(del_sql)
    conn.commit()
    data[['班级','姓名','总分','语文','数学','英语','物理','化学','生物','地理','政治','历史','考试名称','考试日期']].to_sql('scores', con=engine, if_exists='append', index=False)



    




@st.dialog("二次确认")
def sbutton():
    st.warning(f"该考试成绩数据已经存在，请确认是否覆盖")
    ans = st_btn_group(
        buttons=[{"label": "确认", "value": "1", "style": {"color": "red", "font-size": "20px"}},
                  {"label": "取消", "value": "2"}],   
    )  
    if ans == '2':
        st.rerun()
    if ans == '1':
        st.session_state.upload = True
        st.rerun()
        
    # st1, st2 = st.columns(2)
    # if st1.button("确认"):
    #     return True
    # if st2.button("取消"): 
    #     return False
        # del_sql = f'''delete from scores where 考试名称 = '{exam_name}' '''
        
        # engine = create_engine(f'sqlite:///{db_path}', echo=False)
        # data.to_sql('scores', con=engine, if_exists='append', index=False)
        # st.success("Done!")



if uploaded_file is not None:
    cur =conn.cursor()
    wb = openpyxl.load_workbook(uploaded_file)
    sheetnames = wb.sheetnames
    sheet_selector = st.selectbox("Select sheet:",sheetnames)
    df = pd.read_excel(uploaded_file,sheet_selector)
    st.markdown(f"### Currently Selected: `{sheet_selector}`")
    st.write(df)
    exam_name = st.text_input("考试名称")
    df["考试名称"] = exam_name
    date = st.date_input("考试日期", datetime.date.today())
    date_str = date.strftime('%Y-%m-%d')
    df["考试日期"] = date_str
    bt = st.button("上传数据")
    if bt:
        
        # engine = create_engine(f'sqlite:///{db_path}', echo=False)

        # df.to_sql('scores', con=engine, if_exists='append', index=False)
        # st.success("Done!")
        
        # 检查考试名称，并判断是否上传
        test_sql = f'''select count(*) from scores where 考试名称 = '{exam_name}' '''
        # cur = conn.cursor()
        cur.execute(test_sql)
        ans = cur.fetchone()
        
        if ans[0] > 0:
            # st.warning("该考试已经存在，请确认是否覆盖")
            sbutton()
        else: 
            st.session_state.upload = True
    if "upload" in st.session_state and st.session_state.upload:
        delete_insert(cur,df, exam_name)
        st.write("upload Scucess")
        del st.session_state["upload"]

